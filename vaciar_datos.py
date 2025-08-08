import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from copy import copy

# 🧱 Configuración de la app
st.set_page_config(page_title="Vaciar Datos con Estilo", layout="wide")
st.markdown("## 👋 Bienvenido a la herramienta de vaciado de datos")
st.markdown("Sube tus archivos para comenzar el proceso.")
st.divider()

# 📂 Funciones para carga de archivos
def cargar_base(uploaded_file):
    ext = uploaded_file.name.split(".")[-1].lower()
    try:
        if ext == "csv":
            return pd.read_csv(uploaded_file, skiprows=7)
        elif ext == "xlsx":
            return pd.read_excel(uploaded_file, skiprows=7)
    except Exception as e:
        st.error(f"❌ Error al cargar base: {e}")
        return None

def cargar_fuente(uploaded_file):
    ext = uploaded_file.name.split(".")[-1].lower()
    try:
        if ext == "csv":
            return pd.read_csv(uploaded_file)
        elif ext == "xlsx":
            return pd.read_excel(uploaded_file)
    except Exception as e:
        st.error(f"❌ Error al cargar fuente: {e}")
        return None

# 📤 Subida de archivos
archivo_base = st.file_uploader("📂 Archivo base (.csv o .xlsx con estilo)", type=["csv", "xlsx"])
archivo_fuente = st.file_uploader("📥 Archivo fuente (.csv o .xlsx)", type=["csv", "xlsx"])

# 🧪 Depuración
st.write("Archivo base:", archivo_base)
st.write("Archivo fuente:", archivo_fuente)

if not archivo_base or not archivo_fuente:
    st.info("👆 Sube ambos archivos para comenzar.")
    st.stop()

base_df = cargar_base(archivo_base)
fuente_df = cargar_fuente(archivo_fuente)

if base_df is not None and fuente_df is not None:
    if "ID_SOCIEDAD" not in fuente_df.columns:
        st.error("❌ El archivo fuente debe tener la columna 'ID_SOCIEDAD'")
    else:
        filtrado = fuente_df[fuente_df["ID_SOCIEDAD"] == 1]
        st.subheader("✅ Registros filtrados")
        st.dataframe(filtrado)

        if st.button("📄 Generar archivo Excel con estilo"):
            # 🔄 Convertir archivo_base a Excel con estilo original
            buffer_excel = io.BytesIO()
            base_df.to_excel(buffer_excel, index=False, sheet_name="Datos")
            buffer_excel.seek(0)

            wb = load_workbook(buffer_excel)
            ws = wb.active

            # 🎨 Copiar estilos desde encabezado original (fila 1 del archivo original)
            estilo_encabezado = {}
            for col in range(1, ws.max_column + 1):
                celda_original = ws.cell(row=1, column=col)
                estilo_encabezado[col] = {
                    'font': copy(celda_original.font),
                    'fill': copy(celda_original.fill),
                    'alignment': copy(celda_original.alignment),
                    'border': copy(celda_original.border),
                    'number_format': celda_original.number_format
                }

            # 🧼 Limpiar hoja y reescribir encabezado en fila 8
            ws.delete_rows(1, ws.max_row)
            for c_idx, col_name in enumerate(base_df.columns, start=1):
                celda = ws.cell(row=8, column=c_idx, value=col_name)
                estilo = estilo_encabezado.get(c_idx)
                if estilo:
                    celda.font = estilo['font']
                    celda.fill = estilo['fill']
                    celda.alignment = estilo['alignment']
                    celda.border = estilo['border']
                    celda.number_format = estilo['number_format']

            # 📝 Insertar datos desde la fila 9
            for r_idx, row in enumerate(dataframe_to_rows(filtrado, index=False, header=False), start=9):
                for c_idx, value in enumerate(row, start=1):
                    if c_idx < 27:
                        celda = ws.cell(row=r_idx, column=c_idx)
                        if celda.data_type != 'f':
                            celda.value = value

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            st.download_button(
                label="📥 Descargar archivo_final_con_estilo.xlsx",
                data=output,
                file_name="archivo_final_con_estilo.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

